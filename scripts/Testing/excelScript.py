#import openpyxl
#wb = openpyxl.Workbook()
#ws = wb.create_sheet()
#ws.column_dimensions.group('A','D', hidden=True)
#wb.save('F:\ZTEST-TEST\group2.xlsx')
from openpyxl import Workbook
import datetime
wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
#ws['A1'] = 42
# Rows can also be appended
#ws.append([4, 2, 3, 4, 5])
# Python types will automatically be converted
#ws['A3'] = datetime.datetime.now()
# Save the file
#wb.save("F:\ZTEST-TEST\sample.xlsx")
ws['A1'] = 42
#print ("ws['A1']") print (ws['A1'])
ws['A1'] = 48
#ws['E8'] = datetime.datetime.now()
ws['E8'] = 777
count = 0
for a in ws:
    for i in a:
        st = str(i)
        ss = (st[-3:-1])
        ws[ss] = count
        count = count + 1
        if count >= 15:
            continue
ws['A1'] = 48
wb.save("F:\ZTEST-TEST\sample.xlsx")
